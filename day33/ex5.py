# 作者:  Bruce
# 时间： 2022/5/19
import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    # cell = sheet['a1']
    cell = sheet.cell(1, 2)  # 第一行，第二列
    print(cell.value)
    # print(sheet.max_row)

    for row in range(2, sheet.max_row + 1):
        # print(row)
        cell = sheet.cell(row, 3)
        # print(cell.value)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price
    sheet.cell(1, 4).value = 'corrected price'

    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'a5')

    wb.save('t2.xlsx')  # 另起名字保存，不污染原文件
